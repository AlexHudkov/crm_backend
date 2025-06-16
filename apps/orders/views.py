from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.viewsets import ModelViewSet
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from rest_framework.filters import OrderingFilter
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
import openpyxl
from openpyxl.utils import get_column_letter

from .choices import COURSE_TYPE_CHOICES, COURSE_CHOICES, STATUS_CHOICES, COURSE_FORMAT_CHOICES
from .filters import OrderFilter
from .models import Orders, Comment
from .serializers import OrdersSerializer, CommentSerializer


class OrdersViewSet(ModelViewSet):
    queryset = Orders.objects.select_related('manager').all().order_by('-created_at')
    serializer_class = OrdersSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = OrderFilter
    ordering_fields = ["id", "name", "surname", "email", "phone", "age", "course", "course_format", "course_type",
                       "status", "sum", "already_paid", "group", "created_at", "manager"]

    @action(detail=True, methods=['post'], url_path='add-comment')
    def add_comment(self, request, pk=None):
        order = self.get_object()

        if order.manager and order.manager != request.user:
            return Response({"detail": "You are not allowed to comment on this order."}, status=403)

        serializer = CommentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(order=order, author=request.user)

            if not order.manager:
                order.manager = request.user
            if order.status in [None, "New"]:
                order.status = "In Work"
            order.save()

            order_serializer = self.get_serializer(order)
            return Response(order_serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['delete'], url_path='comments/(?P<comment_id>[^/.]+)')
    def delete_comment(self, request, pk=None, comment_id=None):
        comment = get_object_or_404(Comment, id=comment_id)
        if comment.author != request.user:
            return Response(
                {"detail": "You do not have permission to delete this comment."},
                status=status.HTTP_403_FORBIDDEN,
            )
        comment.delete()
        return Response({"detail": "Comment deleted successfully."}, status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='status-options')
    def get_status_options(self, request):
        return Response(dict(STATUS_CHOICES), status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='course-options')
    def get_course_options(self, request):
        return Response(dict(COURSE_CHOICES), status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='course-type-options')
    def get_course_type_options(self, request):
        return Response(dict(COURSE_TYPE_CHOICES), status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='course-format-options')
    def get_course_format_options(self, request):
        return Response(dict(COURSE_FORMAT_CHOICES), status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        filtered_qs = self.filter_queryset(self.get_queryset())

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Orders"

        headers = ["ID", "Name", "Surname", "Email", "Phone", "Age", "Course", "Course Format", "Course Type", "Status",
                   "Created At"]
        sheet.append(headers)

        for order in filtered_qs:
            sheet.append([
                order.id,
                order.name,
                order.surname,
                order.email,
                order.phone,
                order.age,
                order.course,
                order.course_format,
                order.course_type,
                order.status,
                order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else ""
            ])

        for i in range(1, len(headers) + 1):
            col_letter = get_column_letter(i)
            sheet.column_dimensions[col_letter].width = 18

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="orders.xlsx"'
        workbook.save(response)
        return response

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        data = request.data.copy()

        if instance.manager and instance.manager != request.user:
            return Response(
                {"detail": "You are not allowed to edit this order."},
                status=status.HTTP_403_FORBIDDEN
            )

        if data.get("status") == "New":
            data["manager"] = None

        elif (instance.status is None or instance.status == "New") and instance.manager is None:
            data.setdefault("status", "In Work")
            data["manager"] = request.user.id

        serializer = self.get_serializer(instance, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return Response(serializer.data)
